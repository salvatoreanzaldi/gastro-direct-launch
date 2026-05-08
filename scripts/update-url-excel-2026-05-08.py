"""
In-place update der existierenden URL-Liste:
- Quelle:    URL-Liste - Stand 2026-04-28.xlsx
- Ziel:      URL-Liste - Stand 2026-05-08.xlsx
- Änderungen:
  1. Header-Datum auf 2026-05-08
  2. Sheet 1 "Aktuelle URLs": Pro Sprache neuen Sub-Cluster "Vergleiche"
     mit 6 Einträgen (Hub + 5 Detail-Pages) einfügen
  3. Sheet 2 "Geplante URLs": "order-smart"-Einträge auf "Umgesetzt" flippen

Style-Erhalt: Insert + Style-Copy von Referenz-Zellen (Subcluster + Daten).
"""

import copy
import openpyxl

SRC = "URL-Liste - Stand 2026-04-28.xlsx"
DST = "URL-Liste - Stand 2026-05-08.xlsx"
NEW_DATE = "2026-05-08"

VERGLEICHE_SEGMENT = {
    "de": "vergleiche",
    "en": "vs",
    "it": "confronti",
    "fa": "vs",
    "si": "vs",
    "ru": "vs",
}

# Built comparison pages (src/data/comparisons/index.ts)
BUILT_SLUGS = [
    ("order-smart",  "OrderSmart"),
    ("sides",        "SIDES"),
    ("dish-order",   "DISH Order"),
    ("foodamigos",   "Foodamigos"),
    ("resmio",       "resmio"),
]

LANG_SECTION_LABEL = {
    "de": "Vergleiche",
    "en": "Comparisons",
    "it": "Confronti",
    "fa": "Vergleiche",   # Sub-Label im Sheet, nicht User-facing
    "si": "Vergleiche",
    "ru": "Vergleiche",
}

LANG_HUB_NAME = {
    "de": "Vergleiche Hub",
    "en": "Comparisons Hub",
    "it": "Hub Confronti",
    "fa": "Vergleiche Hub",
    "si": "Vergleiche Hub",
    "ru": "Vergleiche Hub",
}

LANG_NOTE = {
    "de": "Konkurrenz-Vergleichsseite, GEO-optimiert",
    "en": "Competitor comparison, GEO-optimized",
    "it": "Pagina di confronto, GEO-optimized",
    "fa": "Konkurrenz-Vergleichsseite, GEO-optimiert",
    "si": "Konkurrenz-Vergleichsseite, GEO-optimiert",
    "ru": "Konkurrenz-Vergleichsseite, GEO-optimiert",
}


def copy_cell_style(src_cell, dst_cell):
    """Style 1:1 von src auf dst übertragen (Font/Fill/Border/Alignment/NumberFormat)."""
    if src_cell.has_style:
        dst_cell.font = copy.copy(src_cell.font)
        dst_cell.fill = copy.copy(src_cell.fill)
        dst_cell.border = copy.copy(src_cell.border)
        dst_cell.alignment = copy.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy.copy(src_cell.protection)


def copy_row_style(ws, src_row, dst_row, max_col):
    for col in range(1, max_col + 1):
        copy_cell_style(ws.cell(row=src_row, column=col), ws.cell(row=dst_row, column=col))
    if src_row in ws.row_dimensions:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


def find_lang_anchors(ws):
    """Returns dict {lang: row_index_of_'▶ Lang'-header}."""
    label_to_lang = {
        "Deutsch (DE)": "de",
        "English (EN)": "en",
        "Italiano (IT)": "it",
        "Farsi (FA)": "fa",
        "Singhalesisch (SI)": "si",
        "Русский (RU)": "ru",
    }
    anchors = {}
    for row in range(1, ws.max_row + 1):
        v = ws.cell(row=row, column=1).value
        if not v:
            continue
        for label, lang in label_to_lang.items():
            if label in str(v):
                anchors[lang] = row
                break
    return anchors


def find_subcluster_reference_rows(ws, lang_row):
    """
    Innerhalb der Sprach-Section: Finde
    - reference subcluster-header row (z.B. '  Legal')
    - reference data rows (alt + non-alt) — typischerweise direkt darunter
    """
    sub_header_row = None
    data_row_alt = None
    data_row_plain = None
    # Suche bis nächste ▶-Zeile oder bis 60 Zeilen weiter
    for row in range(lang_row + 1, min(lang_row + 60, ws.max_row + 1)):
        a = ws.cell(row=row, column=1).value
        b = ws.cell(row=row, column=2).value
        c = ws.cell(row=row, column=3).value
        if a and "▶" in str(a):
            break
        if a and not b and not c and str(a).strip():
            # Subcluster-Header (nur Spalte A gefüllt mit Indent)
            if sub_header_row is None:
                sub_header_row = row
        elif a and b and c:
            # Daten-Zeile
            if data_row_alt is None:
                data_row_alt = row
            elif data_row_plain is None and row != data_row_alt:
                data_row_plain = row
                break
    return sub_header_row, data_row_alt, data_row_plain


def insert_vergleiche_section(ws, insert_at, lang, ref_sub, ref_data_a, ref_data_b, max_col):
    """
    Fügt 7 Zeilen ein an Position insert_at:
    - 1 Subcluster-Header '  Vergleiche'
    - 1 Hub-Eintrag
    - 5 Detail-Einträge
    Stil von ref_sub / ref_data_a / ref_data_b kopieren.
    """
    rows_to_insert = 1 + 1 + len(BUILT_SLUGS)  # = 7

    # FIX: openpyxl.insert_rows shiftet Daten, aber NICHT merged cells.
    # → Vor dem Insert: Merges >= insert_at unmerge, später re-merge an geshifteter Position.
    merges_to_shift = []
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= insert_at:
            merges_to_shift.append(
                (mr.min_row, mr.max_row, mr.min_col, mr.max_col)
            )
            ws.unmerge_cells(start_row=mr.min_row, end_row=mr.max_row,
                             start_column=mr.min_col, end_column=mr.max_col)

    ws.insert_rows(insert_at, amount=rows_to_insert)

    # Re-merge an geshifteter Position
    for (mn, mx, cn, cx) in merges_to_shift:
        ws.merge_cells(
            start_row=mn + rows_to_insert, end_row=mx + rows_to_insert,
            start_column=cn, end_column=cx,
        )

    seg = VERGLEICHE_SEGMENT[lang]
    sub_label = LANG_SECTION_LABEL[lang]
    note = LANG_NOTE[lang]
    hub_name = LANG_HUB_NAME[lang]

    # Row 1: Subcluster-Header (alle Spalten leer außer A1)
    r = insert_at
    copy_row_style(ws, ref_sub, r, max_col)
    # Merge nachbauen falls Original gemerged war
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row == ref_sub and mr.max_row == ref_sub:
            ws.merge_cells(start_row=r, end_row=r,
                           start_column=mr.min_col, end_column=mr.max_col)
            break
    ws.cell(row=r, column=1, value=f"  {sub_label}")

    # Daten-Zeilen
    entries = [(hub_name, f"/{lang}/{seg}", "Hub-Übersicht aller Vergleiche")]
    for slug, name in BUILT_SLUGS:
        entries.append((
            f"{name} Vergleich",
            f"/{lang}/{seg}/{slug}",
            note,
        ))

    for idx, (page_name, url, note_val) in enumerate(entries):
        r = insert_at + 1 + idx
        # Style alternierend: gerade = ref_data_a, ungerade = ref_data_b
        ref = ref_data_a if (idx % 2 == 0) else (ref_data_b or ref_data_a)
        copy_row_style(ws, ref, r, max_col)
        # Werte setzen — Spalten: A=Kategorie, B=Seite, C=URL, D=Status, E=Notizen
        ws.cell(row=r, column=1, value="Vergleiche")
        ws.cell(row=r, column=2, value=page_name)
        ws.cell(row=r, column=3, value=url)
        ws.cell(row=r, column=4, value="Live")
        ws.cell(row=r, column=5, value=note_val)


def update_sheet1(ws):
    """Sheet 1 update: Header-Datum + Vergleiche-Subcluster pro Sprache einfügen."""
    # Header-Datum
    title = ws.cell(row=1, column=1).value
    if title and "Stand:" in str(title):
        new_title = str(title).rsplit("Stand:", 1)[0] + f"Stand: {NEW_DATE}"
        ws.cell(row=1, column=1, value=new_title)

    # Lang-Anker FRISCH ermitteln, dann von unten nach oben einfügen
    # damit Indizes nicht verschieben.
    anchors = find_lang_anchors(ws)
    max_col = ws.max_column

    # Reference rows pro Sprache (von ORIGINAL — vor jeglichem Insert)
    refs = {}
    for lang, lang_row in anchors.items():
        sub, da, db = find_subcluster_reference_rows(ws, lang_row)
        refs[lang] = (sub, da, db)
        if sub is None or da is None:
            print(f"WARN [{lang}] keine Referenz gefunden — Style-Fallback aktiv")

    # Insert-Reihenfolge: von unten nach oben (RU, SI, FA, IT, EN, DE)
    lang_order_bottom_up = ["ru", "si", "fa", "it", "en", "de"]

    for lang in lang_order_bottom_up:
        if lang not in anchors:
            print(f"WARN: Sprache {lang} nicht gefunden, skip")
            continue
        # Section-Endzeile finden (vor nächster ▶-Zeile oder Sheet-Ende)
        anchors_now = find_lang_anchors(ws)
        lang_row = anchors_now[lang]
        # Suche nächste ▶-Zeile
        next_lang_row = ws.max_row + 1
        for other_lang, row in anchors_now.items():
            if other_lang != lang and row > lang_row and row < next_lang_row:
                next_lang_row = row
        insert_at = next_lang_row  # vor nächste Sprache (oder Sheet-Ende)

        ref_sub, ref_da, ref_db = refs[lang]
        if ref_sub is None:
            ref_sub = lang_row + 1  # Fallback
        if ref_da is None:
            ref_da = lang_row + 2
        if ref_db is None:
            ref_db = ref_da

        insert_vergleiche_section(ws, insert_at, lang, ref_sub, ref_da, ref_db, max_col)
        added = 1 + 1 + len(BUILT_SLUGS)
        print(f"[Sheet1] {lang}: {added} Zeilen eingefügt @ Row {insert_at}")


def update_sheet2(ws):
    """Sheet 2 update: Header-Datum + 'order-smart'-Einträge auf Umgesetzt."""
    title = ws.cell(row=1, column=1).value
    if title and "Stand:" in str(title):
        new_title = str(title).rsplit("Stand:", 1)[0] + f"Stand: {NEW_DATE}"
        ws.cell(row=1, column=1, value=new_title)

    flipped = 0
    for row in range(1, ws.max_row + 1):
        url = ws.cell(row=row, column=3).value
        if not url:
            continue
        url_s = str(url).strip()
        # Nur die "order-smart"-Einträge — andere bleiben Geplant
        # (URL-Patterns: /de/vergleiche/order-smart, /en/comparisons/order-smart,
        #  /it/confronti/order-smart, /fa/comparisons/order-smart, etc.)
        if url_s.endswith("/order-smart") and any(
            seg in url_s for seg in ["vergleiche", "comparisons", "confronti"]
        ):
            ws.cell(row=row, column=4, value="Umgesetzt")
            # Spalte 5 ist Priorität; ergänze Hinweis-Marker via Notiz an Spalte 5 nicht überschreiben
            # → stattdessen URL-Spalte ist klar; ergänze Hinweis in der freien letzten Spalte falls vorhanden
            # Sheet 2 hat 5 Spalten (laut Struktur): kein Notes-Slot.
            # → Wir setzen den Hinweis in eine 6. Spalte (sicherer Ansatz: Spalte F).
            ws.cell(row=row, column=6, value="Live — siehe Sheet 1")
            flipped += 1

    print(f"[Sheet2] {flipped} Einträge auf 'Umgesetzt' geflippt")
    return flipped


def main():
    wb = openpyxl.load_workbook(SRC)
    ws1 = wb["Aktuelle URLs"]
    ws2 = wb["Geplante URLs"]

    print(f"=== Loading {SRC} ===")
    print(f"Sheet1 vor Update: {ws1.max_row} Zeilen")
    print(f"Sheet2 vor Update: {ws2.max_row} Zeilen")

    update_sheet1(ws1)
    flipped = update_sheet2(ws2)

    wb.save(DST)
    print(f"\n=== Saved: {DST} ===")
    print(f"Sheet1 nach Update: {ws1.max_row} Zeilen")
    print(f"Sheet2 nach Update: {ws2.max_row} Zeilen")
    print(f"Sheet2 'Umgesetzt' geflippt: {flipped}")


if __name__ == "__main__":
    main()
