"""
Generate "Gastro Master - Komplette URL-Struktur 2026-04-21.xlsx"
Sheet 1: Aktuelle URLs (alle Live-Seiten, 6 Sprachen)
Sheet 2: Geplante URLs (geplante Seiten, nur DE)
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ─── Farben ───────────────────────────────────────────────────────────────────
CLR_HEADER_BG   = "0A264A"   # Dunkelblau
CLR_HEADER_FG   = "FFFFFF"
CLR_LIVE        = "D6F0D3"   # Hellgrün
CLR_LIVE_FG     = "1A6B1A"
CLR_GEPLANT     = "FFF8CC"   # Gelb
CLR_GEPLANT_FG  = "7A5C00"
CLR_SPAETER     = "E8E8E8"   # Grau
CLR_SPAETER_FG  = "555555"
CLR_PRIO_H      = "FFD7D7"   # Rot
CLR_PRIO_H_FG   = "8B0000"
CLR_PRIO_M      = "FFE8CC"   # Orange
CLR_PRIO_M_FG   = "7A3800"
CLR_PRIO_L      = "E8E8E8"   # Grau
CLR_PRIO_L_FG   = "555555"
CLR_ROW_ALT     = "F7F9FC"   # Zeilen-Alternation

LANGUAGES = ["de", "en", "it", "fa", "si", "ru"]
LANG_LABELS = {
    "de": "Deutsch (DE)",
    "en": "English (EN)",
    "it": "Italiano (IT)",
    "fa": "فارسی (FA)",
    "si": "Slovenščina (SI)",
    "ru": "Русский (RU)",
}

# ─── Aktuelle Seiten (aus routes.ts) ─────────────────────────────────────────
# (Kategorie, Seite, Pfad, Notiz)
LIVE_PAGES = [
    # Start
    ("Start", "Homepage", "/", "Hauptseite"),

    # Produkte / Pakete
    ("Produkte / Pakete", "Produkte Übersicht", "/produkte", "Produkt-Hub"),
    ("Produkte / Pakete", "Online-Bestellshop", "/produkte/pakete/online-bestellshop", "Webshop Paket"),
    ("Produkte / Pakete", "Bestell-App", "/produkte/pakete/bestell-app", "White-Label App"),
    ("Produkte / Pakete", "Webseite", "/produkte/pakete/webseite", "Restaurant-Webseite"),
    ("Produkte / Pakete", "Kassensystem", "/produkte/pakete/kassensystem", "POS-System"),

    # Produkte / Add-Ons
    ("Produkte / Add-Ons", "Add-Ons Übersicht", "/produkte/add-ons", "Add-On Hub"),
    ("Produkte / Add-Ons", "QR-Code Flyer", "/produkte/add-ons/qr-code-flyer", "Tischbestellung via QR"),
    ("Produkte / Add-Ons", "Fahrer-App GPS", "/produkte/add-ons/fahrer-app-gps", "GPS-Tracking Fahrer"),
    ("Produkte / Add-Ons", "QR-Code Tischsystem", "/produkte/add-ons/qr-code-tischsystem", "Tischsystem QR"),
    ("Produkte / Add-Ons", "Bildschirmfunktion", "/produkte/add-ons/bildschirmfunktion", "Küchendisplay"),
    ("Produkte / Add-Ons", "Kiosk", "/produkte/add-ons/kiosk", "Self-Service Kiosk"),
    ("Produkte / Add-Ons", "Transaktionsumlage", "/produkte/add-ons/transaktionsumlage", "Gebühren-Weitergabe"),

    # Produkte / Hardware
    ("Produkte / Hardware", "Hardware Übersicht", "/produkte/hardware", "Hardware-Katalog"),

    # Lösungen
    ("Lösungen", "Lösungen Übersicht", "/loesungen", "Solutions Hub"),
    ("Lösungen", "Lieferservice Gründen", "/loesungen/lieferservice-gruenden", "Step-by-Step Guide"),
    ("Lösungen", "Franchise", "/loesungen/franchise", "Franchise-System"),
    ("Lösungen", "Restaurant", "/loesungen/restaurant", "Für Restaurants"),
    ("Lösungen", "Lieferdienst", "/loesungen/lieferdienst", "Für Lieferdienste"),
    ("Lösungen", "Café & Bäckerei", "/loesungen/cafe-baeckerei", "Für Cafés/Bäckereien"),
    ("Lösungen", "Ghost Kitchen", "/loesungen/ghost-kitchen", "Dark Kitchen"),

    # Blog
    ("Blog", "Blog Übersicht", "/blog", "Blog-Hub"),
    ("Blog", "Warum Lieferando verzichten?", "/blog/warum-lieferando-verzichten", "SEO-Artikel"),
    ("Blog", "5 Fehler Lieferdienst eröffnen", "/blog/5-fehler-lieferdienst-eroffnen", "SEO-Artikel"),
    ("Blog", "Was kostet ein Bestellsystem?", "/blog/was-kostet-bestellsystem", "SEO-Artikel"),

    # Info
    ("Info", "Preise", "/preise", "Preisübersicht"),
    ("Info", "FAQ", "/faq", "Häufige Fragen"),
    ("Info", "Über uns", "/uber-uns", "Team & Mission"),
    ("Info", "Integrationen", "/integrations", "Alle Integrationen"),
    ("Info", "Kontakt", "/kontakt", "Kontaktformular"),

    # Downloads
    ("Downloads", "Downloads Übersicht", "/downloads", "Software-Downloads"),
    ("Downloads", "Druckertreiber", "/downloads/druckertreiber", "Epson Treiber"),

    # Legal
    ("Legal", "Impressum", "/impressum", ""),
    ("Legal", "Datenschutz", "/datenschutz", "DSGVO"),
    ("Legal", "AGB", "/agb", "Allg. Geschäftsbedingungen"),
]

# ─── Geplante Seiten ──────────────────────────────────────────────────────────
# (Cluster, Seite, Pfad, Priorität, Notiz)
PLANNED_PAGES = [
    # Integrationen
    ("Integrationen", "Lieferando Integration", "/de/integrationen/lieferando", "HOCH", "Bestellsync Lieferando"),
    ("Integrationen", "Wolt Integration", "/de/integrationen/wolt", "HOCH", "Delivery Partner"),
    ("Integrationen", "Uber Eats Integration", "/de/integrationen/uber-eats", "HOCH", "Uber Eats Sync"),
    ("Integrationen", "Stripe Integration", "/de/integrationen/stripe", "HOCH", "Payment Gateway"),
    ("Integrationen", "Klarna Integration", "/de/integrationen/klarna", "MITTEL", "Buy now, pay later"),
    ("Integrationen", "Mollie Integration", "/de/integrationen/mollie", "MITTEL", "Zahlungsabwicklung"),
    ("Integrationen", "Adyen Integration", "/de/integrationen/adyen", "MITTEL", "Enterprise Payments"),
    ("Integrationen", "VISA Zahlung", "/de/integrationen/visa", "MITTEL", "Kartenzahlung"),
    ("Integrationen", "Mastercard Zahlung", "/de/integrationen/mastercard", "MITTEL", "Kartenzahlung"),
    ("Integrationen", "Datev Integration", "/de/integrationen/datev", "MITTEL", "Buchaltungsexport"),
    ("Integrationen", "TSE Kassensicherung", "/de/integrationen/tse", "HOCH", "GoBD-Konformität"),
    ("Integrationen", "Epson Drucker", "/de/integrationen/epson", "NIEDRIG", "Bondrucker"),
    ("Integrationen", "DHL Versand", "/de/integrationen/dhl", "NIEDRIG", "Paketversand"),
    ("Integrationen", "DPD Versand", "/de/integrationen/dpd", "NIEDRIG", "Paketversand"),
    ("Integrationen", "UPS Versand", "/de/integrationen/ups", "NIEDRIG", "Paketversand"),
    ("Integrationen", "WhatsApp Business", "/de/integrationen/whatsapp", "MITTEL", "Kundenkommunikation"),
    ("Integrationen", "Google Integration", "/de/integrationen/google", "HOCH", "Google Maps / My Business"),
    ("Integrationen", "Facebook Integration", "/de/integrationen/facebook", "MITTEL", "Social Commerce"),
    ("Integrationen", "Instagram Shopping", "/de/integrationen/instagram", "MITTEL", "Social Commerce"),
    ("Integrationen", "Make (Integromat)", "/de/integrationen/make", "NIEDRIG", "Automation"),

    # Wissen – Lieferdienst gründen
    ("Wissen", "Lieferdienst gründen (Anleitung)", "/de/wissen/lieferdienst-gruenden", "HOCH", "Step-by-Step Guide"),
    ("Wissen", "Was kostet ein Lieferdienst?", "/de/wissen/lieferdienst-kosten", "HOCH", "Kostenübersicht"),
    ("Wissen", "Marketing für Lieferdienste", "/de/wissen/lieferdienst-marketing", "MITTEL", "Best Practices"),

    # Wissen – Bestellsystem
    ("Wissen", "Bestellsystem Vergleich", "/de/wissen/bestellsystem-vergleich", "HOCH", "Feature Comparison"),
    ("Wissen", "Bestellsystem richtig auswählen", "/de/wissen/bestellsystem-auswahl", "HOCH", "Selection Guide"),
    ("Wissen", "Digitale Gastronomie Basics", "/de/wissen/digitale-gastronomie", "MITTEL", "Digitaler Wandel"),

    # Wissen – App erstellen
    ("Wissen", "Eigene App erstellen (Guide)", "/de/wissen/app-erstellen", "HOCH", "White-Label App Guide"),
    ("Wissen", "Was kostet eine Bestellapp?", "/de/wissen/app-kosten", "HOCH", "Cost Calculator"),
    ("Wissen", "App oder Webshop?", "/de/wissen/app-vs-webshop", "MITTEL", "Entscheidungshilfe"),

    # Wissen – Webshop
    ("Wissen", "Webshop für Restaurants", "/de/wissen/webshop-gastronomie", "HOCH", "Setup Guide"),
    ("Wissen", "Webshop für Lieferdienste", "/de/wissen/webshop-lieferdienste", "HOCH", "Specialized Guide"),

    # Ratgeber
    ("Ratgeber", "Kundenbindung Strategien", "/de/ratgeber/kundenbindung-restaurant", "MITTEL", "Retention Tips"),
    ("Ratgeber", "Online Umsatz steigern", "/de/ratgeber/online-umsatz-steigern", "MITTEL", "Sales Optimization"),
    ("Ratgeber", "TSE Konformität Kassensystem", "/de/ratgeber/tse-kassensystem", "MITTEL", "Legal Compliance"),
    ("Ratgeber", "Lieferdienst Prozesse optimieren", "/de/ratgeber/lieferdienst-prozesse", "NIEDRIG", "Operations Guide"),
    ("Ratgeber", "Personal Management Tipps", "/de/ratgeber/personal-management", "NIEDRIG", "HR Guide"),

    # Vergleiche
    ("Vergleiche", "Order Smart Vergleich", "/de/vergleiche/order-smart", "HOCH", "Detaillierte Gegenüberstellung"),
    ("Vergleiche", "App Smart Vergleich", "/de/vergleiche/app-smart", "HOCH", "Feature Comparison"),
    ("Vergleiche", "Lieferando Alternative", "/de/vergleiche/lieferando-alternative", "HOCH", "SEO: Alternativen"),
    ("Vergleiche", "Wolt Alternative", "/de/vergleiche/wolt-alternative", "MITTEL", "SEO: Alternativen"),

    # Regional (später)
    ("Regional (Geplant)", "Bestellsystem Frankfurt", "/de/frankfurt/bestellsystem", "NIEDRIG", "Regional Targeting"),
    ("Regional (Geplant)", "Lieferservice Berlin", "/de/berlin/lieferservice", "NIEDRIG", "Regional Targeting"),
    ("Regional (Geplant)", "Kassensystem Hamburg", "/de/hamburg/kassensystem", "NIEDRIG", "Regional Targeting"),
    ("Regional (Geplant)", "Bestellsystem München", "/de/muenchen/bestellsystem", "NIEDRIG", "Regional Targeting"),
    ("Regional (Geplant)", "Lieferservice Köln", "/de/koeln/lieferservice", "NIEDRIG", "Regional Targeting"),
    ("Regional (Geplant)", "Kassensystem Stuttgart", "/de/stuttgart/kassensystem", "NIEDRIG", "Regional Targeting"),
    ("Regional (Geplant)", "Lieferservice Düsseldorf", "/de/duesseldorf/lieferservice", "NIEDRIG", "Regional Targeting"),
    ("Regional (Geplant)", "Bestellsystem Leipzig", "/de/leipzig/bestellsystem", "NIEDRIG", "Regional Targeting"),
]


def make_fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def make_border():
    side = Side(style="thin", color="D0D7E3")
    return Border(left=side, right=side, top=side, bottom=side)


def style_header_row(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = make_fill(CLR_HEADER_BG)
        cell.font = Font(bold=True, color=CLR_HEADER_FG, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = make_border()


def style_section_header(ws, row, cols, label):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1)
    cell.value = label
    cell.fill = make_fill("1B4B8A")
    cell.font = Font(bold=True, color="FFFFFF", size=11, italic=True)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cell.border = make_border()


# ─── SHEET 1: Aktuelle URLs ───────────────────────────────────────────────────
def build_sheet1(wb):
    ws = wb.active
    ws.title = "Aktuelle URLs"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    # Row 1: Titel
    ws.merge_cells("A1:E1")
    title = ws.cell(row=1, column=1, value="Gastro Master — Aktuelle URLs (Live) | Stand: 2026-04-21")
    title.fill = make_fill("0A264A")
    title.font = Font(bold=True, color="FFFFFF", size=13)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Row 2: Header
    headers = ["Kategorie", "Seite / Produktname", "URL (/de/...)", "Status", "Notizen"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    style_header_row(ws, 2, 5)
    ws.row_dimensions[2].height = 22

    current_row = 3
    current_lang = None
    current_cat  = None
    row_idx      = 0

    for lang in LANGUAGES:
        # Sprach-Trennzeile
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
        lang_cell = ws.cell(row=current_row, column=1,
                            value=f"  ▶  {LANG_LABELS[lang]}")
        lang_cell.fill = make_fill("163D6E")
        lang_cell.font = Font(bold=True, color="FFFFFF", size=11)
        lang_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[current_row].height = 20
        current_row += 1
        current_cat = None

        for (cat, page, path, note) in LIVE_PAGES:
            # Kategorie-Gruppe
            if cat != current_cat:
                style_section_header(ws, current_row, 5, f"  {cat}")
                ws.row_dimensions[current_row].height = 18
                current_row += 1
                current_cat = cat
                row_idx = 0

            full_url = f"/{lang}{path}" if path != "/" else f"/{lang}/"
            bg = CLR_ROW_ALT if row_idx % 2 == 0 else "FFFFFF"

            data = [cat, page, full_url, "Live", note]
            for col, val in enumerate(data, 1):
                cell = ws.cell(row=current_row, column=col, value=val)
                cell.fill = make_fill(bg)
                cell.border = make_border()
                cell.alignment = Alignment(vertical="center", wrap_text=(col == 3))
                cell.font = Font(size=10)

            # Status-Zelle
            status_cell = ws.cell(row=current_row, column=4)
            status_cell.fill = make_fill(CLR_LIVE)
            status_cell.font = Font(bold=True, color=CLR_LIVE_FG, size=10)
            status_cell.alignment = Alignment(horizontal="center", vertical="center")

            ws.row_dimensions[current_row].height = 16
            current_row += 1
            row_idx += 1

    # Spaltenbreiten
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 52
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 34

    return ws


# ─── SHEET 2: Geplante URLs ───────────────────────────────────────────────────
def build_sheet2(wb):
    ws = wb.create_sheet(title="Geplante URLs")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    # Row 1: Titel
    ws.merge_cells("A1:F1")
    title = ws.cell(row=1, column=1,
                    value="Gastro Master — Geplante URLs (Roadmap) | Stand: 2026-04-21")
    title.fill = make_fill("0A264A")
    title.font = Font(bold=True, color="FFFFFF", size=13)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Row 2: Header
    headers = ["Cluster / Kategorie", "Seite / Thema", "URL (/de/...)", "Status", "Priorität", "Notizen / Grund"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    style_header_row(ws, 2, 6)
    ws.row_dimensions[2].height = 22

    current_row = 3
    current_cluster = None
    row_idx = 0

    STATUS_MAP = {
        "Regional (Geplant)": ("Geplant (Später)", CLR_SPAETER, CLR_SPAETER_FG),
    }

    for (cluster, page, path, prio, note) in PLANNED_PAGES:
        # Cluster-Trennzeile
        if cluster != current_cluster:
            style_section_header(ws, current_row, 6, f"  {cluster}")
            ws.row_dimensions[current_row].height = 18
            current_row += 1
            current_cluster = cluster
            row_idx = 0

        bg = CLR_ROW_ALT if row_idx % 2 == 0 else "FFFFFF"

        # Status
        if cluster in STATUS_MAP:
            status_text, status_bg, status_fg = STATUS_MAP[cluster]
        else:
            status_text, status_bg, status_fg = "Geplant", CLR_GEPLANT, CLR_GEPLANT_FG

        # Priorität
        prio_map = {
            "HOCH":   (CLR_PRIO_H, CLR_PRIO_H_FG),
            "MITTEL": (CLR_PRIO_M, CLR_PRIO_M_FG),
            "NIEDRIG":(CLR_PRIO_L, CLR_PRIO_L_FG),
        }
        prio_bg, prio_fg = prio_map.get(prio, (CLR_PRIO_L, CLR_PRIO_L_FG))

        data = [cluster, page, path, status_text, prio, note]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=current_row, column=col, value=val)
            cell.fill = make_fill(bg)
            cell.border = make_border()
            cell.alignment = Alignment(vertical="center", wrap_text=(col == 3))
            cell.font = Font(size=10)

        ws.cell(row=current_row, column=4).fill = make_fill(status_bg)
        ws.cell(row=current_row, column=4).font = Font(bold=True, color=status_fg, size=10)
        ws.cell(row=current_row, column=4).alignment = Alignment(horizontal="center", vertical="center")

        ws.cell(row=current_row, column=5).fill = make_fill(prio_bg)
        ws.cell(row=current_row, column=5).font = Font(bold=True, color=prio_fg, size=10)
        ws.cell(row=current_row, column=5).alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[current_row].height = 16
        current_row += 1
        row_idx += 1

    # Spaltenbreiten
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 34

    return ws


# ─── Main ─────────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
build_sheet1(wb)
build_sheet2(wb)

output_path = "/Users/salvatore/Desktop/Gastro Master - Komplette URL-Struktur 2026-04-21.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")

# Stats
live_rows = len(LIVE_PAGES) * len(LANGUAGES)
planned_rows = len(PLANNED_PAGES)
print(f"Sheet 1 — Aktuelle URLs: {live_rows} Zeilen ({len(LIVE_PAGES)} Seiten × {len(LANGUAGES)} Sprachen)")
print(f"Sheet 2 — Geplante URLs: {planned_rows} Zeilen")
